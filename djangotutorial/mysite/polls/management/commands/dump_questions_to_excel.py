from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Write questions to an excel file"

    def handle(self, *args, **options):
        alphabet = string.ascii_lowercase
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Question"
        for column in range(1, 6):
            ws[alphabet[column] + "1"] = "Choice " + str(column)
        for num, question in enumerate(Question.objects.all()):
            ws[alphabet[0] + str(num + 2)] = question.question_text
            for num2, choice in enumerate(question.choice_set.all()):
                ws[alphabet[num2 + 1] + str(num + 2)]= choice.choice_text
        wb.save("questions_dump.xlsx")
