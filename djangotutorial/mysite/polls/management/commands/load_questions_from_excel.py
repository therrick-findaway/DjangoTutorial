import sys
from polls.models import Question
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook


def import_questions_and_choices_workbook(file):
    return load_workbook(filename=file)


def check_for_multiple_sheets(wb):
    if len(wb.worksheets) > 1:
        raise Exception('Too many worksheets. Make sure there is only one.')


def return_sheet_data_from_workbook(wb):
    return wb.worksheets[0]


def generate_data_from_excel_sheet(sheet):
    for num in sheet.iter_rows(min_row=2, values_only=True):
        yield num


def ingest_question_and_choices(question_and_choices):
    new_question = Question(question_text=question_and_choices[0], pub_date=timezone.now())
    new_question.save()
    for i in range(1, len(question_and_choices)):
        if question_and_choices[i] is not None:
            new_question.choice_set.create(choice_text=question_and_choices[i])
        else:
            break


class Command(BaseCommand):
    help = 'Read questions from excel file to database'

    def add_arguments(self, parser):
        parser.add_argument('filename', type=str)

    def handle(self, *args, **options):
        with transaction.atomic(): 
            workbook = import_questions_and_choices_workbook(options['filename'])
            check_for_multiple_sheets(workbook)
            sheet_data = return_sheet_data_from_workbook(workbook)
            for question_and_choices in generate_data_from_excel_sheet(sheet_data):
                ingest_question_and_choices(question_and_choices)
            print('Questions and choices successfully uploaded.')    
